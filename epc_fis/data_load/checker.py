import pandas as pd
import numpy as np
import argparse
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.sql_tools import SqlConnection
from utils.constants.load_constants import SqlTables

#Limpiar hardcoding!!!!

def arg_parser():
    parser = argparse.ArgumentParser(description="Chequeo de discrepancias en PCRs")
    parser.add_argument('--run', help = 'Nombre del Run que se quiere comprobar',
                        required=True)
    args = parser.parse_args()
    return args


def main():
    now = datetime.now().strftime('%Y-%m-%d_%H:%M:%S')
    args = arg_parser()
    run = args.run
    sql_constants = SqlTables()
    sql_connect = SqlConnection(sql_constants.schema)

    #variables:
    pcr_suffix = '_pcr'
    neg = 'neg'
    pos= 'pos'
    obs_list = ['OXA', 'VIM', 'KPC', 'NDM', 'CBP', 'ANT']
    #run = '1ªPCR_16052023'
    print(f"Script iniciado en: {now}")
    print(f'Comprobando run {run}')
    print('Consultando base de datos')

    pcr = pd.read_sql(f'SELECT * FROM {sql_constants.pcr_general} WHERE run = "{run}"', sql_connect.engine)#Hay que meter los pd.read_sql en SqlConnection
 
    if pcr.empty:
        sys.exit(f"No existe el run indicado: {run}")

    results = pd.read_sql(
        f"""
        SELECT {sql_constants.sample_name}, {sql_constants.cod_mo} 
        from {sql_constants.sil_results}""", sql_connect.engine)

    obs = pd.read_sql(f"""
                SELECT {sql_constants.sample_name}, {sql_constants.obs} 
                from {sql_constants.sil_obs}""", 
                sql_connect.engine).drop_duplicates()

    pcr = pcr[pcr.columns.drop(list(pcr.filter(regex='origin_file')))]

    pcr[sql_constants.cq] = pcr[sql_constants.cq].fillna(0)

    levels = pcr[sql_constants.target].unique()

    if len(levels) == 1 and levels[0] == sql_constants.val_screening:
        print('Consultando PCR de cribado')
    elif len(levels) >1 and sql_constants.val_screening not in levels:
        print("Consultando PCR específica")
    else:
        print("Consultando PCR con esquema no esperado. Se debe revisar con atención la salida")

    pcr = pcr[[sql_constants.sample, 
               sql_constants.run, sql_constants.target, sql_constants.cq]
          ].pivot_table(index=[sql_constants.sample, sql_constants.run], 
                        values=sql_constants.cq, columns=sql_constants.target
                        ).add_suffix(pcr_suffix).reset_index() 

    print('Comenzando limpieza de datos')
    pcr_col_levels = [f'{value}{pcr_suffix}' for value in levels if value != f'16S{pcr_suffix}']

    obs = obs[obs[sql_constants.obs].isin(obs_list)]
    obs['Values'] = pos
    obs = obs.pivot(index=sql_constants.sample_name, columns=sql_constants.obs, 
                    values='Values').fillna(neg)

    #Cruzamos resultados con observaciones

    results[sql_constants.cod_mo] = results[sql_constants.cod_mo].mask(
        results[sql_constants.cod_mo] != neg, pos)

    cross_df = results.merge(obs, how='outer', on=sql_constants.sample_name, indicator=True)

    #Arreglamos el cross_df:

    cross_df = cross_df.drop_duplicates()
    cross_df[obs_list] = cross_df[obs_list].fillna(neg)

    cross_df = cross_df.rename(columns={sql_constants.cod_mo : "culture"})

    cross_df = cross_df.drop('_merge', axis=1)

    # Tienes que cruzar con results para obtener los negativos!

    print('Cruzando datos!')
    check_file = pcr.merge(cross_df, how='left', left_on=sql_constants.sample, right_on=sql_constants.sample_name, indicator=True)

    #Chekeos complementarios
    #Buscamos señales de carba:
    print('Comprobando discrepancias!')
    check_file['carbapenemase'] = check_file[obs_list].apply(
    lambda x: pos if any(x == pos) else neg, axis=1
    )

#Creamos patrón de colores
    cond_data_missing = check_file['_merge'] == 'left_only'

    target_result = check_file[pcr_col_levels].apply(
    lambda x: any(x == 0), axis=1
    )

    cond_data_disc = target_result & (check_file['carbapenemase'] == pos)

    check_file = check_file.drop('_merge', axis=1)
    check_file['disc'] = cond_data_disc

    print('Escribiendo archivo excel!')
    wb = Workbook()

    ws = wb.active

    for r in dataframe_to_rows(check_file, index=False, header=True):
        ws.append(r)

    for i, valor in enumerate(cond_data_disc, start=2):
        if valor:
            for cell in ws[i]:
                cell.fill = PatternFill(start_color='00FFFF00', fill_type='solid') #aMARILLO


    for i, valor in enumerate(cond_data_missing, start=2):
        if valor:
            for cell in ws[i]:
                cell.fill = PatternFill(start_color='00FF0000', fill_type='solid') #Rojo

    wb.save(f'consulta_run-{run}_{now}.xlsx')
    print('Proceso finalizado!')